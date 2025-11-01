import requests
import pandas as pd
from datetime import datetime, date, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional

# --- Constants for Configuration & Styling ---

# API Configuration
API_URL = "https://shreejiattendance.run.place/api/attendance/logs"
API_TIMEOUT = 15  # seconds

# --- ADD SECRETS HERE ---
# These must match the secrets on your ESP32 device
ESP_SECRET = ""
SERVER_SECRET = ""	

# File & Salary Configuration
OUTPUT_FILENAME = "Pivoted_Attendance_Report_with_Salary.xlsx"
COMPANY_NAME = "Shreeji Remedies"
REPORT_TITLE = "Monthly Attendance & Salary Report"
FULL_DAY_HOURS = 10.5 # Define how many hours constitute a full working day

# --- Define Employee Financial Data ---
# Use employee 'uid' as the key. 'default' is a fallback value.
EMPLOYEE_MONTHLY_SALARIES = {
    'B3517733': 13000, 'B3166E33': 15000, 'E3967F33': 14000, '10': 16000, 'default': 12000
}
EMPLOYEE_ALLOWANCES = {'default': 0}
EMPLOYEE_ADVANCES = {'default': 0}
EMPLOYEE_LOANS = {'default': 0}
EMPLOYEE_PREMIUMS = {'default': 0}


# --- STYLES ---

# Fills
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Slate Blue
SUBHEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid") # Lighter Slate Blue
CI_HEADER_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid") # Light Green
CO_HEADER_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Light Red
MINS_HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid") # Light Blue
SUNDAY_FILL = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid") # Light Red for Sunday
TODAY_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Light Yellow
ALT_COL_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid") # Light Grey/Blue for columns
FINANCIAL_HEADER_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid") # Orange for Financial Summary
DEDUCTIONS_HEADER_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid") # Green for Deductions
# New styles from image
NAME_CELL_FILL = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid") # Light Red/Orange
IN_HAND_SALARY_FILL = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid") # Muted Green

# Fonts
HEADER_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
TABLE_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUNDAY_DATE_FONT = Font(name="Calibri", size=11, bold=True, color="943126") # Dark Red
TODAY_DATE_FONT = Font(name="Calibri", size=11, bold=True, color="B7950B") # Dark Yellow/Gold
CI_FONT = Font(name="Calibri", color="287431", bold=True) # Dark Green
CO_FONT = Font(name="Calibri", color="943126", bold=True) # Dark Red
MINS_FONT = Font(name="Calibri", color="1B4F72", bold=True) # Dark Blue
DATA_FONT = Font(name="Calibri", size=11)
CI_DATA_FONT = Font(name="Calibri", size=11, color="287431")
CO_DATA_FONT = Font(name="Calibri", size=11, color="943126")
IN_HAND_SALARY_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="145A32") # Dark Green for contrast

# Borders
THIN_BORDER_SIDE = Side(style="thin", color="BDC3C7") # Grey
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

# Alignment
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# --- Token Generation ---
def generate_auth_token(esp_secret: str, server_secret: str) -> str:
    if not esp_secret or not server_secret or "YOUR_" in esp_secret or "YOUR_" in server_secret:
        print("⚠️ [Auth] Secrets are not set. Skipping token generation.")
        return ""
    token_parts = [f'{(ord(e) ^ ord(server_secret[i % len(server_secret)])):02x}' for i, e in enumerate(esp_secret)]
    return "".join(token_parts)

# --- API Data Fetching ---
def fetch_attendance_data(url: str) -> Dict[str, Any]:
    try:
        token = generate_auth_token(ESP_SECRET, SERVER_SECRET)
        headers = {'x-esp32-token': token} if token else {}
        if token: print(f"🔐 [Auth] Using token: {token}")
        response = requests.get(url, headers=headers, timeout=API_TIMEOUT)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data from API: {e}\nContinuing with an empty dataset.")
        return {'data': []}

# --- Data Processing ---
def get_month_dates(year: int, month: int) -> List[datetime]:
    try:
        start_date = datetime(year, month, 1)
        days_in_month = pd.Period(f'{year}-{month}').days_in_month
        return [start_date + pd.Timedelta(days=i) for i in range(days_in_month)]
    except ValueError:
        print(f"❌ Invalid year ({year}) or month ({month}). Exiting.")
        return []

def process_payload(payload: Dict[str, Any], dates: List[datetime]) -> pd.DataFrame:
    attendance_list = []
    for emp in payload.get('data', []):
        log_map = {dt.date(): [] for dt in dates}
        for log in emp.get('logs', []):
            try:
                log_date = datetime.strptime(log['date'], "%d/%m/%Y").date()
                if log_date in log_map:
                    log_map[log_date].append(log)
            except (ValueError, KeyError):
                continue
        row = {"No.": emp.get("uid", "N/A"), "Name": emp.get("name", "Unknown"), **log_map}
        attendance_list.append(row)
    df = pd.DataFrame(attendance_list)
    return df if df.empty else df.sort_values(by="No.")

def map_financial_data(df: pd.DataFrame) -> pd.DataFrame:
    """Maps the static financial data (salaries, allowances, etc.) to each employee."""
    if df.empty: return df
    
    def map_data(data_dict):
        return [data_dict.get(str(id), data_dict.get('default', 0)) for id in df['No.']]

    df['monthly_salary'] = map_data(EMPLOYEE_MONTHLY_SALARIES)
    df['allowance'] = map_data(EMPLOYEE_ALLOWANCES)
    df['advance_paid'] = map_data(EMPLOYEE_ADVANCES)
    df['loan'] = map_data(EMPLOYEE_LOANS)
    df['premium'] = map_data(EMPLOYEE_PREMIUMS)
    
    return df

# --- Excel Sheet Creation ---
def style_cell(cell, font: Font, alignment: Alignment, fill: Optional[PatternFill] = None, border: Optional[Border] = None):
    cell.font, cell.alignment = font, alignment
    if fill: cell.fill = fill
    if border: cell.border = border

def create_report_header(ws, month_name: str, year: int):
    ws.merge_cells("A1:F2")
    ws["A1"].value = f"{COMPANY_NAME}\n{REPORT_TITLE} - {month_name} {year}"
    style_cell(ws["A1"], HEADER_FONT, CENTER_ALIGN, HEADER_FILL)
    ws.merge_cells("G1:J2")
    ws["G1"].value = f"Report Generated:\n{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    style_cell(ws["G1"], SUBTITLE_FONT, CENTER_ALIGN, HEADER_FILL)

def create_table_headers(ws, df: pd.DataFrame):
    ws.merge_cells("A4:A5"), ws.merge_cells("B4:B5")
    ws["A4"].value, ws["B4"].value = "Date", "Status"
    style_cell(ws["A4"], TABLE_HEADER_FONT, CENTER_ALIGN, SUBHEADER_FILL, THIN_BORDER)
    style_cell(ws["B4"], TABLE_HEADER_FONT, CENTER_ALIGN, SUBHEADER_FILL, THIN_BORDER)
    if df.empty: return
    for i, (_, emp_row) in enumerate(df.iterrows()):
        col = 3 + i
        ws.cell(row=4, column=col, value=emp_row["No."])
        ws.cell(row=5, column=col, value=emp_row["Name"])
        style_cell(ws.cell(row=4, column=col), TABLE_HEADER_FONT, CENTER_ALIGN, SUBHEADER_FILL, THIN_BORDER)
        style_cell(ws.cell(row=5, column=col), TABLE_HEADER_FONT, CENTER_ALIGN, SUBHEADER_FILL, THIN_BORDER)

# ---
# --- THIS IS THE UPDATED FUNCTION ---
# ---
def populate_attendance_data(ws, df: pd.DataFrame, dates: List[datetime], today: date):
    start_row = 6
    FULL_DAY_MINUTES = FULL_DAY_HOURS * 60
    RELIEF_MINUTES = 30
    
    for i, dt in enumerate(dates):
        row_ci, row_co, row_mins = start_row + (i * 3), start_row + (i * 3) + 1, start_row + (i * 3) + 2
        
        is_sunday, is_today = dt.weekday() == 6, dt.date() == today
        day_fill, date_font = (SUNDAY_FILL, SUNDAY_DATE_FONT) if is_sunday else ((TODAY_FILL, TODAY_DATE_FONT) if is_today else (None, DATA_FONT))

        ws.merge_cells(start_row=row_ci, start_column=1, end_row=row_mins, end_column=1)
        ws.cell(row_ci, 1, f"{dt.strftime('%a')}\n{dt.day}")
        for r in range(row_ci, row_mins + 1): style_cell(ws.cell(r, 1), date_font, CENTER_ALIGN, day_fill, THIN_BORDER)
        
        ws.cell(row_ci, 2, "Check-In"), ws.cell(row_co, 2, "Check-Out"), ws.cell(row_mins, 2, "Minutes Worked")
        style_cell(ws.cell(row_ci, 2), CI_FONT, CENTER_ALIGN, CI_HEADER_FILL, THIN_BORDER)
        style_cell(ws.cell(row_co, 2), CO_FONT, CENTER_ALIGN, CO_HEADER_FILL, THIN_BORDER)
        style_cell(ws.cell(row_mins, 2), MINS_FONT, CENTER_ALIGN, MINS_HEADER_FILL, THIN_BORDER)

        if df.empty: continue
        for j, (_, emp_row) in enumerate(df.iterrows()):
            col = 3 + j
            
            # --- MODIFIED BLOCK: Parse time strings into time objects ---
            
            logs = emp_row.get(dt.date(), [])
            
            # Note: The time format from your API is likely "%I:%M:%S %p" (e.g., 09:57:55 am)
            # We must convert this string to a real time object for Excel.
            time_format = "%I:%M:%S %p"  
            checkins_dt = []
            checkouts_dt = []
            
            for l in logs:
                try:
                    # Parse the time string into a datetime.time object
                    time_str = l['time'].lower().strip() # Clean up string
                    time_obj = datetime.strptime(time_str, time_format).time()
                    if l['status'] == "Check-IN":
                        checkins_dt.append(time_obj)
                    elif l['status'] == "Check-OUT":
                        checkouts_dt.append(time_obj)
                except (ValueError, KeyError):
                    continue  # Skip logs with bad time format or missing keys

            # Now find the min/max of the actual time objects, not strings
            ci_val = min(checkins_dt) if checkins_dt else "-"
            co_val = max(checkouts_dt) if checkouts_dt else "-"
            
            ci_cell, co_cell = ws.cell(row_ci, col, ci_val), ws.cell(row_co, col, co_val)
            
            # When writing a time object, explicitly set the number format
            if isinstance(ci_val, time):
                ci_cell.number_format = 'h:mm:ss AM/PM'
            if isinstance(co_val, time):
                co_cell.number_format = 'h:mm:ss AM/PM'
                
            ci_ref, co_ref = ci_cell.coordinate, co_cell.coordinate

            # --- END OF MODIFIED BLOCK ---

            actual_minutes_formula = f'ROUND(({co_ref}-{ci_ref})*1440,2)'
            
            # This formula is now correct and will work with the numeric time values
            minutes_formula = (
                f'=IF(OR({ci_ref}="-",{co_ref}="-",{co_ref}<{ci_ref}), 0, '
                f'LET(ActualMins, {actual_minutes_formula}, ExpectedMins, {FULL_DAY_MINUTES}, '
                f'IF(ActualMins >= ExpectedMins, ActualMins, '
                f'LET(Shortfall, ExpectedMins - ActualMins, '
                f'IF(Shortfall <= {RELIEF_MINUTES}, ExpectedMins, ActualMins)'
                f'))))'
            )

            if is_sunday:
                half_day_minutes = FULL_DAY_MINUTES / 2
                minutes_formula = (f'=IF(OR({ci_ref}="-",{co_ref}="-",{co_ref}<{ci_ref}),0,'
                                 f'IF({actual_minutes_formula}>={half_day_minutes},{FULL_DAY_MINUTES},{actual_minutes_formula}))')
            
            mins_cell = ws.cell(row_mins, col, minutes_formula)
            mins_cell.number_format = '#,##0.00'
            
            col_fill = ALT_COL_FILL if j % 2 != 0 else None
            style_cell(ci_cell, CI_DATA_FONT, CENTER_ALIGN, col_fill, THIN_BORDER)
            style_cell(co_cell, CO_DATA_FONT, CENTER_ALIGN, col_fill, THIN_BORDER)
            style_cell(mins_cell, DATA_FONT, CENTER_ALIGN, col_fill, THIN_BORDER)
            
        ws.row_dimensions[row_ci].height, ws.row_dimensions[row_co].height, ws.row_dimensions[row_mins].height = 20, 20, 20


def create_pivoted_summary(ws, df: pd.DataFrame, month_dates: List[datetime], start_col: int):
    start_row = 4
    ws.cell(start_row, start_col, "Financial Summary")
    style_cell(ws.cell(start_row, start_col), HEADER_FONT, CENTER_ALIGN, FINANCIAL_HEADER_FILL)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(df))

    # Row Headers (Metrics)
    headers = ["ID", "Name", "Presence", "Absence", "Basic Salary", "Total Days", "Payable Days", "Per Day Amt",
               "Per Min Wage", "Total Mins", "Gross Salary", "Short Hour Deduct.", "Earned Salary", "In Hand Salary"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_row + 2 + i, column=start_col)
        cell.value = h
        # Special styling for the 'In Hand Salary' header
        if h == "In Hand Salary":
            style_cell(cell, IN_HAND_SALARY_HEADER_FONT, Alignment(horizontal="right"), IN_HAND_SALARY_FILL, THIN_BORDER)
        else:
            style_cell(cell, TABLE_HEADER_FONT, Alignment(horizontal="right"), SUBHEADER_FILL, THIN_BORDER)

    if df.empty: return

    num_dates = len(month_dates)
    
    # Column Headers (Employees) and Data
    for i, (_, emp_row) in enumerate(df.iterrows()):
        current_col = start_col + 1 + i
        
        # --- Formula Generation ---
        main_table_col_letter = get_column_letter(3 + i)
        full_range = f"{main_table_col_letter}6:{main_table_col_letter}{5 + (num_dates * 3)}"
        start_cell_full_range = f"{main_table_col_letter}6"
        
        # Helper to get current column's cell coordinates
        def cc(row_idx): return get_column_letter(current_col) + str(start_row + 2 + row_idx)

        # Static Data (ID & Name) are handled in the styling loop below
        
        # Formulas
        ws.cell(row=start_row + 4, column=current_col, value=f'=SUMPRODUCT(--(MOD(ROW({full_range})-ROW({start_cell_full_range}),3)=2),--({full_range}>0))') # Presence
        ws.cell(row=start_row + 5, column=current_col, value=f"={cc(5)}-{cc(2)}") # Absence
        ws.cell(row=start_row + 6, column=current_col, value=emp_row['monthly_salary']).number_format = '"₹"#,##0.00'
        ws.cell(row=start_row + 7, column=current_col, value=num_dates) # Total Days
        ws.cell(row=start_row + 8, column=current_col, value=f"={cc(2)}") # Payable Days
        ws.cell(row=start_row + 9, column=current_col, value=f"=IF({cc(5)}>0,{cc(4)}/{cc(5)},0)").number_format = '"₹"#,##0.00'
        ws.cell(row=start_row + 10, column=current_col, value=f"=IF({cc(5)}>0,{cc(4)}/({cc(5)}*{FULL_DAY_HOURS}*60),0)").number_format = '"₹"#,##0.0000'
        ws.cell(row=start_row + 11, column=current_col, value=f'=SUMPRODUCT(--(MOD(ROW({full_range})-ROW({start_cell_full_range}),3)=2),{full_range})').number_format = '#,##0.00'
        ws.cell(row=start_row + 12, column=current_col, value=f"={cc(7)}*{cc(6)}").number_format = '"₹"#,##0.00'
        ws.cell(row=start_row + 13, column=current_col, value=f"=MAX(0, ({cc(6)}*{FULL_DAY_HOURS}*60 - {cc(9)})*{cc(8)})").number_format = '"₹"#,##0.00'
        ws.cell(row=start_row + 14, column=current_col, value=f"={cc(10)}-{cc(11)}").number_format = '"₹"#,##0.00'
        
        deductions_table_start_col = start_col + len(df) + 2  
        allowance_ref = get_column_letter(deductions_table_start_col + 2) + str(start_row + 3 + i)
        deductions_total_ref = get_column_letter(deductions_table_start_col + 6) + str(start_row + 3 + i)
        ws.cell(row=start_row + 15, column=current_col, value=f"={cc(12)}+{allowance_ref}-{deductions_total_ref}").number_format = '"₹"#,##0.00'

        # --- Styling Loop ---
        for r_idx in range(2, 16): # Loop through all metric rows
            cell = ws.cell(row=start_row + r_idx, column=current_col)
            fill = None # Default to no fill
            
            # Static data for ID and Name
            if r_idx == 2: cell.value = emp_row["No."]
            if r_idx == 3: 
                cell.value = emp_row["Name"]
                fill = NAME_CELL_FILL
            
            # Special fill for In Hand Salary
            if r_idx == 15:
                fill = IN_HAND_SALARY_FILL

            style_cell(cell, DATA_FONT, CENTER_ALIGN, fill, THIN_BORDER)


def create_deductions_table(ws, df: pd.DataFrame, start_col: int):
    start_row = 4
    ws.cell(start_row, start_col, "Allowances & Deductions")
    style_cell(ws.cell(start_row, start_col), HEADER_FONT, CENTER_ALIGN, DEDUCTIONS_HEADER_FILL)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 6)
    
    headers = ["ID", "Name", "Allowance", "Advance Paid", "Loan", "Premium", "Total Deductions"]
    header_row = start_row + 2
    for i, h in enumerate(headers):
        style_cell(ws.cell(header_row, start_col + i, h), TABLE_HEADER_FONT, CENTER_ALIGN, SUBHEADER_FILL, THIN_BORDER)

    if df.empty: return
    top_data_row = header_row + 1
    for i, (_, emp_row) in enumerate(df.iterrows()):
        r = top_data_row + i
        ws.cell(r, start_col + 0, emp_row['No.'])
        ws.cell(r, start_col + 1, emp_row['Name'])
        ws.cell(r, start_col + 2, emp_row['allowance']).number_format = '"₹"#,##0.00'
        ws.cell(r, start_col + 3, emp_row['advance_paid']).number_format = '"₹"#,##0.00'
        ws.cell(r, start_col + 4, emp_row['loan']).number_format = '"₹"#,##0.00'
        ws.cell(r, start_col + 5, emp_row['premium']).number_format = '"₹"#,##0.00'
        
        c1 = get_column_letter(start_col + 3)
        c2 = get_column_letter(start_col + 5)
        ws.cell(r, start_col + 6, f"=SUM({c1}{r}:{c2}{r})").number_format = '"₹"#,##0.00'

        for c_idx in range(7):
            style_cell(ws.cell(r, start_col + c_idx), DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)

def finalize_styles(ws, num_main_cols: int):
    # Main table
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 10, 15
    for i in range(3, num_main_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    # Dynamically set width for all summary columns
    for i in range(num_main_cols + 1, num_main_cols + 30):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = get_column_letter(3) + "6"
    ws.sheet_view.show_grid_lines = False

# --- Main Execution ---
def get_user_date_input() -> (Optional[int], Optional[int]):
    today = datetime.now()
    default_year, default_month = today.year, today.month
    try:
        year_in = input(f"Enter year [Enter for {default_year}]: ")
        year = int(year_in) if year_in else default_year
        month_in = input(f"Enter month (1-12) [Enter for {default_month}]: ")
        month = int(month_in) if month_in else default_month
        if not 1 <= month <= 12: month = default_month
        return year, month
    except ValueError:
        return default_year, default_month

def main():
    print("🚀 Starting attendance report generation...")
    year, month = get_user_date_input()
    if not (year and month): return
    month_dates = get_month_dates(year, month)
    if not month_dates: return
    
    month_name, today_date = month_dates[0].strftime("%B"), datetime.now().date()
    print(f"Fetching data for {month_name} {year}...")
    payload = fetch_attendance_data(API_URL)
    df = process_payload(payload, month_dates)
    
    if df.empty: print("⚠️ No employee data. Report will have headers only.")
    else: df = map_financial_data(df)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {month_name} {year}"
    
    print("Building Excel workbook...")
    create_report_header(ws, month_name, year)
    create_table_headers(ws, df)
    
    num_main_cols = 2 + len(df) if not df.empty else 2
    summary_start_col = num_main_cols + 2 # Add a gap column
    create_pivoted_summary(ws, df, month_dates, summary_start_col)

    deductions_start_col = summary_start_col + len(df) + 2 if not df.empty else summary_start_col + 2
    create_deductions_table(ws, df, deductions_start_col)
    
    populate_attendance_data(ws, df, month_dates, today_date)

    finalize_styles(ws, num_main_cols)

    try:
        wb.save(OUTPUT_FILENAME)
        print(f"\n✅ Success! Report saved as '{OUTPUT_FILENAME}'")
    except IOError as e:
        print(f"\n❌ Error saving file: {e}")

if __name__ == "__main__":
    main()
